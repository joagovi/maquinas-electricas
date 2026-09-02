#!/usr/bin/env python3
"""Genera el Excel de inscripcion a las exposiciones del Vibequest.

Marca como cerradas las sesiones que ya pasaron y abre cupos solo en las que
faltan, para que nadie se inscriba a una fecha que ya ocurrio.

El alumno escribe su codigo y el nombre aparece solo; si se inscribe dos veces
la celda se pone roja. La hoja "Alumnos" indica quien falta por inscribirse.

Subelo a Drive COMO HOJA DE GOOGLE, no lo mandes adjunto: si lo adjuntas al
correo recibiras una copia distinta por alumno.

Uso:
    .venv/bin/python scripts/inscripcion.py [salida.xlsx]
"""

from __future__ import annotations

import datetime
import sys
from pathlib import Path

import yaml
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

RAIZ = Path(__file__).resolve().parent.parent
DIAS = ["lunes", "martes", "miércoles", "jueves", "viernes", "sábado", "domingo"]
MESES = ["enero", "febrero", "marzo", "abril", "mayo", "junio", "julio",
         "agosto", "setiembre", "octubre", "noviembre", "diciembre"]

AZUL = PatternFill("solid", fgColor="1F4E79")
GRIS = PatternFill("solid", fgColor="D9D9D9")
VERDE = PatternFill("solid", fgColor="C6EFCE")
AMBAR = PatternFill("solid", fgColor="FFF2CC")
ROJO = PatternFill("solid", fgColor="FFC7CE")
BORDE = Border(*[Side(style="thin", color="BFBFBF")] * 4)


def fecha_de(s) -> datetime.date:
    f = s["fecha"]
    return f if isinstance(f, datetime.date) else datetime.date.fromisoformat(str(f))


def bonito(d: datetime.date) -> str:
    return f"{DIAS[d.weekday()]} {d.day} de {MESES[d.month - 1]}"


def main() -> int:
    curso = yaml.safe_load((RAIZ / "silabo" / "curso.yml").read_text(encoding="utf-8"))
    interno = yaml.safe_load((RAIZ / "silabo" / "interno.yml").read_text(encoding="utf-8"))
    alumnos = interno["alumnos"]
    cfg = curso["exposiciones"]
    por = cfg["alumnos_por_sesion"]
    hoy = datetime.date.today()

    temas = {c.get("semana"): c["titulo"] for c in curso.get("capitulos", []) if c.get("semana")}
    ses = [s for s in curso["sesiones"]
           if s["tipo"] == "clase" and s.get("responsable") == cfg.get("responsable")]

    abiertas = [s for s in ses if fecha_de(s) >= hoy]
    cupos = len(abiertas) * por

    wb = Workbook()
    ws = wb.active
    ws.title = "Inscripción"
    for col, ancho in zip("ABCDEF", (5, 13, 38, 40, 40, 26)):
        ws.column_dimensions[col].width = ancho

    ws["A1"] = f'Vibequest — inscripción de exposiciones · {curso["codigo"]} {curso["ciclo"]}'
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws.merge_cells("A1:F1")

    reglas = [
        "Escribe tu CÓDIGO en la columna B; el nombre aparece solo. Elige la FECHA que te convenga:",
        f"el tema es el de esa semana. Hay {por} cupos por sesión, {cfg['minutos_por_alumno']} minutos por persona más preguntas.",
        f"Exponer SUMA hasta {cfg['bonus']['maximo']} puntos sobre la práctica calificada de esa unidad. No exponer no penaliza.",
        "Si tu código sale en ROJO es que está repetido: ya te inscribiste en otra sesión.",
        "Pega el enlace a tu Colab y a la conversación con la IA. Ambos deben quedar VISIBLES "
        "('cualquiera con el enlace') hasta que se publique la nota.",
        "Edita SOLO tu fila. Todo cambio queda registrado en el historial de versiones.",
    ]
    for i, r in enumerate(reglas, start=2):
        ws.cell(row=i, column=1, value=r).font = Font(
            italic=True, size=9, color="C00000" if "ROJO" in r else "404040")
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=6)

    fila = 7
    celdas = []
    for s in ses:
        f = fecha_de(s)
        pasada = f < hoy
        n = s["n"]
        ws.cell(row=fila, column=1, value=f"Sesión {n}")
        ws.cell(row=fila, column=2, value=bonito(f))
        ws.cell(row=fila, column=3, value=temas.get(n, s.get("tema") or "por definir"))
        ya = s.get("expusieron")
        if pasada:
            etiqueta = "YA SE DICTÓ — sin exposiciones" if ya == 0 else f"YA SE DICTÓ — {ya} expusieron"
        else:
            etiqueta = f"{por} cupos"
        ws.cell(row=fila, column=4, value=etiqueta)
        ws.cell(row=fila, column=5, value="")
        for col in range(1, 7):
            cc = ws.cell(row=fila, column=col)
            cc.fill = GRIS if pasada else AZUL
            cc.font = Font(bold=True, color="404040" if pasada else "FFFFFF")
        fila += 1

        encabezados = ["#", "Código", "Nombre (se completa solo)",
                       "Enlace a tu Colab", "Enlace a la conversación con IA", ""]
        for col, h in enumerate(encabezados, 1):
            cc = ws.cell(row=fila, column=col, value=h)
            cc.font = Font(bold=True, size=9)
            cc.fill = AMBAR
            cc.border = BORDE
        fila += 1

        n_filas = (s.get("expusieron") or 0) if pasada else por
        if pasada and n_filas == 0:
            fila += 1
            continue
        for k in range(1, n_filas + 1):
            ws.cell(row=fila, column=1, value=k).border = BORDE
            cel = ws.cell(row=fila, column=2)
            cel.border = BORDE
            if pasada:
                cel.fill = VERDE
                ws.cell(row=fila, column=6, value="← anotar quién expuso").font = Font(
                    italic=True, size=8, color="808080")
            celdas.append((f"B{fila}", pasada))
            form = ws.cell(row=fila, column=3)
            form.value = (f'=IFERROR(VLOOKUP(B{fila},Alumnos!$A:$B,2,FALSE),'
                          f'IF(B{fila}="","","código no encontrado"))')
            form.border = BORDE
            form.font = Font(size=10)
            for col in (4, 5):
                ws.cell(row=fila, column=col).border = BORDE
            fila += 1
        fila += 1

    wa = wb.create_sheet("Alumnos")
    for col, ancho in zip("ABC", (12, 46, 18)):
        wa.column_dimensions[col].width = ancho
    for col, h in enumerate(["Código", "Nombre", "¿Ya se inscribió?"], 1):
        cc = wa.cell(row=1, column=col, value=h)
        cc.font = Font(color="FFFFFF", bold=True)
        cc.fill = AZUL
    for i, a in enumerate(alumnos, start=2):
        wa.cell(row=i, column=1, value=int(a["codigo"])).number_format = "0"
        wa.cell(row=i, column=2, value=a["nombre"])
        wa.cell(row=i, column=3,
                value=f'=IF(COUNTIF(Inscripción!$B:$B,A{i})>0,"sí","— falta")')
    ultima = len(alumnos) + 1
    wa.conditional_formatting.add(f"C2:C{ultima}", FormulaRule(formula=['$C2="— falta"'], fill=AMBAR))
    wa.conditional_formatting.add(f"C2:C{ultima}", FormulaRule(formula=['$C2="sí"'], fill=VERDE))

    dv = DataValidation(type="list", formula1=f"=Alumnos!$A$2:$A${ultima}", allow_blank=True,
                        showErrorMessage=True, errorTitle="Código no válido",
                        error="Elige tu código de la lista de matriculados.")
    ws.add_data_validation(dv)
    for cd, _ in celdas:
        dv.add(ws[cd])
        ws.conditional_formatting.add(cd, FormulaRule(
            formula=[f'AND({cd}<>"",COUNTIF(Inscripción!$B:$B,{cd})>1)'], fill=ROJO))

    ws.freeze_panes = "A7"
    salida = Path(sys.argv[1]) if len(sys.argv) > 1 else RAIZ / "Vibequest - inscripcion.xlsx"
    wb.save(salida)

    print(f"[ok] {salida}")
    print(f"     {len(ses)} sesiones; {len(abiertas)} abiertas, {len(ses)-len(abiertas)} ya dictadas")
    print(f"     {cupos} cupos disponibles para {len(alumnos)} alumnos")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
