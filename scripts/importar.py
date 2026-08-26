#!/usr/bin/env python3
"""Importa la informacion interna del curso a silabo/interno.yml.

Lee dos Excel directamente del Google Drive montado (~/cursos/drive-pucp) y los
deja en un YAML manejable:

  .../Alumnos *.xlsx     lista de matriculados
  .../HORARIO *.xlsx     horarios de clase, practica y laboratorio, con
                         profesores y jefes de practica

No se copia nada al disco: el Drive se lee bajo demanda. Si el Drive no esta
montado, se usan como respaldo _entrada/listas/ y silabo/.

De los alumnos extrae SOLO codigo, nombre y horario. Descarta correos, seguros y
cualquier otra columna: no hacen falta y no queremos esos datos circulando.

El archivo que produce, silabo/interno.yml, esta FUERA DE GIT a proposito:
contiene nombres de alumnos y de jefes de practica.

Uso:
    .venv/bin/python scripts/importar.py
"""

from __future__ import annotations

import atexit
import os
import re
import subprocess
import sys
import tempfile
import unicodedata
from pathlib import Path

try:
    import openpyxl
    import yaml
except ImportError:
    sys.exit("Faltan dependencias.  Ejecuta:  .venv/bin/pip install openpyxl pyyaml")

RAIZ = Path(__file__).resolve().parent.parent
LISTAS = RAIZ / "_entrada" / "listas"
SALIDA = RAIZ / "silabo" / "interno.yml"

# Google Drive montado con rclone (servicio drive-pucp.service). El material
# vive ahi: no se duplica nada en el disco.
DRIVE = Path.home() / "cursos" / "drive-pucp"
CICLO = "2026-2"
ROOT_ID = "1U28ia49D2zIV0FsbQNgNq0T8zjkrsfZz"

# Temporales con datos de alumnos: se borran siempre al terminar, pase lo que pase.
_TEMPORALES: list[Path] = []


@atexit.register
def _limpiar_temporales() -> None:
    for t in _TEMPORALES:
        try:
            t.unlink(missing_ok=True)
        except OSError:
            pass


def buscar_en_drive(patron: str) -> list[Path]:
    """Busca por nombre dentro de la carpeta del ciclo en el Drive montado."""
    base = DRIVE / CICLO
    if not base.is_dir():
        return []
    return sorted(
        p
        for p in base.rglob(patron)
        if p.is_file() and not p.name.startswith("~$")
    )


def localizar(patron: str, carpeta_local: Path) -> list[Path]:
    """Drive primero; si no esta montado, la carpeta local de respaldo."""
    if encontrados := buscar_en_drive(patron):
        return encontrados
    if carpeta_local.is_dir():
        return sorted(p for p in carpeta_local.glob(patron) if not p.name.startswith("~$"))
    return []


def materializar(ruta: Path) -> Path:
    """Devuelve una ruta legible de verdad.

    Algunos archivos del Drive fueron editados en Google Sheets y quedan con
    tamano desconocido: el montaje los muestra pero al leerlos devuelve 0 bytes.
    En ese caso se descargan a un temporal con `rclone cat`, que si funciona.
    """
    if ruta.stat().st_size > 0:
        return ruta

    relativa = ruta.relative_to(DRIVE)
    # Temporal con permisos 600 y borrado al terminar: estos archivos llevan
    # nombres de alumnos y no deben quedarse en /tmp.
    fd, nombre_tmp = tempfile.mkstemp(prefix="importar-", suffix=ruta.suffix)
    os.close(fd)
    destino = Path(nombre_tmp)
    destino.chmod(0o600)
    _TEMPORALES.append(destino)
    print(f"    (0 bytes por el montaje; descargando con rclone: {ruta.name})")

    with destino.open("wb") as fh:
        res = subprocess.run(
            [
                str(Path.home() / ".local" / "bin" / "rclone"),
                "cat",
                f"pucp:{relativa.as_posix()}",
                "--drive-root-folder-id",
                ROOT_ID,
            ],
            stdout=fh,
            stderr=subprocess.PIPE,
            check=False,
        )
    if res.returncode != 0 or destino.stat().st_size == 0:
        raise RuntimeError(
            f"No pude leer {ruta.name} ni por el montaje ni con rclone.\n"
            f"    {res.stderr.decode('utf-8', 'ignore')[:200]}"
        )
    return destino

RE_CODIGO = re.compile(r"^\s*(\d{8})(?:\.0)?\s*$")

TIPOS = {"cla": "clase", "pra": "practica", "lab": "laboratorio", "exa": "examen"}


def limpiar(valor) -> str:
    return " ".join(str(valor).split()) if valor is not None else ""


def entero(valor) -> int | None:
    try:
        return int(float(valor))
    except (TypeError, ValueError):
        return None


def sin_tildes(texto: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFD", texto) if unicodedata.category(c) != "Mn"
    )


def lineas(valor) -> list[str]:
    """Varias personas o sesiones vienen separadas por saltos de linea."""
    return [" ".join(p.split()) for p in str(valor or "").split("\n") if p.strip()]


# "SAB 15:00-18:00 C N422"  ->  dia, horas, secuencia C, aula N422
# La letra suelta antes del aula es la SECUENCIA, no parte del nombre del aula:
#   C = clase,  A y B = las dos secuencias que alternan cada semana (por eso
#   practica y laboratorio son quincenales).
RE_SESION = re.compile(
    r"^(?P<dia>[A-ZÁÉÍÓÚ]{3})\s+"
    r"(?P<horas>\d{1,2}:\d{2}\s*-\s*\d{1,2}:\d{2})\s+"
    r"(?P<secuencia>[A-Z])\s+"
    r"(?P<aula>\S+)\s*$"
)


def partir_sesion(texto: str) -> dict:
    """Separa 'SAB 15:00-18:00 C N422' en sus componentes."""
    m = RE_SESION.match(texto.strip())
    if not m:
        return {"crudo": texto}
    d = m.groupdict()
    return {
        "dia": d["dia"],
        "horas": d["horas"].replace(" ", ""),
        "secuencia": d["secuencia"],
        "aula": d["aula"],
    }


# --------------------------------------------------------------------------- #
def fila_encabezado(filas: list[tuple], marcador: str) -> int:
    for i, fila in enumerate(filas):
        if any(marcador in limpiar(c) for c in fila):
            return i
    raise LookupError(f"No encuentro la columna '{marcador}' en el Excel.")


def importar_horarios(ruta: Path) -> dict:
    ws = openpyxl.load_workbook(materializar(ruta), data_only=True).worksheets[0]
    filas = list(ws.iter_rows(values_only=True))

    i_enc = fila_encabezado(filas, "Tipo Hor.")
    enc = [limpiar(c) for c in filas[i_enc]]
    col = {nombre: enc.index(nombre) for nombre in enc if nombre}

    def valor(fila, nombre):
        return fila[col[nombre]] if nombre in col else None

    salida: dict[str, list] = {}
    for fila in filas[i_enc + 1 :]:
        tipo_bruto = limpiar(valor(fila, "Tipo Hor.")).lower()
        if not tipo_bruto:
            continue
        tipo = TIPOS.get(tipo_bruto[:3], tipo_bruto)

        entrada = {
            "codigo": entero(valor(fila, "Hor.")),
            "vacantes": entero(valor(fila, "Vac.")),
            "matriculados": entero(valor(fila, "Mat.")),
            "sesiones": [partir_sesion(s) for s in lineas(valor(fila, "Sesiones"))],
            "personal": lineas(valor(fila, "Profesor")),
        }
        salida.setdefault(tipo, []).append(entrada)
    return salida


def importar_alumnos(ruta: Path) -> list[dict]:
    ws = openpyxl.load_workbook(materializar(ruta), data_only=True).worksheets[0]
    filas = list(ws.iter_rows(values_only=True))

    try:
        i_enc = fila_encabezado(filas, "Alumno")
        enc = [limpiar(c) for c in filas[i_enc]]
        col = {n: enc.index(n) for n in enc if n}
        i_cod = col.get("Alumno")
        i_nom = col.get("Nombre")
        i_hor = col.get("Horario")
    except LookupError:
        i_enc, i_cod, i_nom, i_hor = -1, None, None, None

    alumnos: dict[str, dict] = {}
    for fila in filas[i_enc + 1 :]:
        codigo = nombre = horario = None

        if i_cod is not None and (m := RE_CODIGO.match(limpiar(fila[i_cod]))):
            codigo = m.group(1)
            nombre = limpiar(fila[i_nom]) if i_nom is not None else None
            horario = entero(fila[i_hor]) if i_hor is not None else None
        else:  # sin encabezado reconocible: buscar por forma
            for celda in fila:
                texto = limpiar(celda)
                if codigo is None and (m := RE_CODIGO.match(texto)):
                    codigo = m.group(1)
                elif nombre is None and len(texto) > 6 and re.search(r"[A-Za-z]{3}", texto):
                    if "@" not in texto:
                        nombre = texto

        if codigo and nombre:
            alumnos.setdefault(codigo, {"codigo": codigo, "nombre": nombre, "horario": horario})

    return sorted(alumnos.values(), key=lambda a: sin_tildes(a["nombre"]))


# --------------------------------------------------------------------------- #
def main() -> int:
    datos: dict = {}

    if DRIVE.is_dir():
        print(f"[i] Leyendo del Drive montado: {DRIVE / CICLO}")
    else:
        print(f"[!] Drive NO montado en {DRIVE}. Uso las carpetas locales.")
        print("    Para montarlo:  systemctl --user start drive-pucp.service")

    horarios_xlsx = localizar("*HORARIO*.xlsx", RAIZ / "silabo")
    if horarios_xlsx:
        datos["horarios"] = importar_horarios(horarios_xlsx[0])
        print(f"[ok] {horarios_xlsx[0].name}")
        for tipo, items in datos["horarios"].items():
            print(f"       {tipo:12} {len(items)} horario(s)")
    else:
        print("[!] No hay Excel de horarios en silabo/.")

    listas = localizar("*Alumnos*.xlsx", LISTAS)
    if listas:
        datos["alumnos"] = importar_alumnos(listas[0])
        print(f"[ok] {listas[0].name}: {len(datos['alumnos'])} alumnos")
    else:
        print(f"[!] No hay Excel de alumnos en {LISTAS.relative_to(RAIZ)}/.")

    if not datos:
        sys.exit("No habia nada que importar.")

    SALIDA.write_text(
        "# Generado por scripts/importar.py.  NO EDITAR A MANO.\n"
        "# FUERA DE GIT a proposito: contiene nombres de alumnos y de jefes de practica.\n\n"
        + yaml.safe_dump(datos, allow_unicode=True, sort_keys=False),
        encoding="utf-8",
    )
    print(f"\n[ok] {SALIDA.relative_to(RAIZ)} escrito (fuera de git).")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
